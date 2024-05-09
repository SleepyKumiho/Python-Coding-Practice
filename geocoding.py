import googlemaps
import openpyxl
import csv

api_key = ""

def geocode_from_address(address, key = api_key):
    """
    Use the googlemaps API to geocode an address (turn a physical address into latitude and longitude)

    Input:
    address (str): A human readable address
    key (str): A Google Maps Platform API key, defaults to a hardcoded variable or can be entered at runtime

    Return:
    latitude, longitude: Tuple of floats represeting the decimal latitude and longitude, or None if it could not be found
    """
    gmaps = googlemaps.Client(key = key)
    geocode_result = gmaps.geocode(address)

    # If the address cannot be geocoded, just return None 
    try:
        latitude = geocode_result[0]["geometry"]["location"]["lat"]
        longitude = geocode_result[0]["geometry"]["location"]["lng"]
        return latitude, longitude
    except:
        return None, None

def geocode_file_excel(input_file_name, output_file_name, key = api_key):
    """
    Opens an Excel file containing a column of addresses (the full addresses should be in the rightmost column)
    Each address is geocoded using the Google Maps Platform API, and the latitude and longitude are written into a new 
    output file. 

    Input:
    input_file_name (str): The path to the input file 
    output_file_name (str): The path to the output file (will overwrite files with the same name)
    key (str): Defaults to a hardcoded value, or can be entered at runtime. A Google Maps Platform API key. 

    Output:
    int: 1 on a succesful completion.
    """
    input_wb = openpyxl.load_workbook(filename = input_file_name)
    input_ws = input_wb.active
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active

    for row_num, row in enumerate(input_ws.values):

        address = row[-1]
        # The first row is colulmn names
        if row_num == 0:
            output_ws.cell(row = row_num+1, column = 1, value = address)
            output_ws.cell(row = row_num+1, column = 2, value = "Latitude")
            output_ws.cell(row = row_num+1, column = 3, value = "Longitude")
            print(address, "lat", "long")
            continue

        lat, long = geocode_from_address(address, key)
        print(address, lat, long)
        output_ws.cell(row = row_num+1, column = 1, value = address)
        output_ws.cell(row = row_num + 1, column = 2, value = lat)
        output_ws.cell(row = row_num + 1, column = 3, value = long)

    output_wb.save(output_file_name)

    return 1

def geocode_file_csv(input_file_name, output_file_name, key = api_key):
    None


if __name__ == '__main__':
    #address = "Jeremy Ranch Park & Ride, Park City, UT"
    #print(geocode_from_address(address))
    if api_key == "":
        key = input("Please enter your Google Maps Platform API key:\n")
        print("For future reference, the API key can be hardcoded in the file to prevent this message.")
        geocode_file_excel("../../../Desktop/test.xlsx", "../../../Desktop/output.xlsx", key)
    else:
        geocode_file_excel("../../../Desktop/test.xlsx", "../../../Desktop/output.xlsx")
